import pandas as pd
import os
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re  # Para eliminar los ** de las celdas

# Configuración para NVIDIA llama-3.1
client = OpenAI(
    base_url="https://integrate.api.nvidia.com/v1",
    api_key="nvapi" # Colocar API KEY
)



def convertir_excel_a_texto(file_path):
    """
    Convierte un archivo Excel en texto procesable.
    Genera un archivo de texto temporal con el contenido ordenado, respetando los formatos de hora originales del archivo.
    """
    try:
        # Cargar el archivo Excel
        df = pd.read_excel(file_path)
        print("Archivo Excel cargado con éxito.")
    except Exception as e:
        print(f"Error al cargar el archivo Excel: {e}")
        return None

    columnas = df.columns.tolist()
    print(f"Columnas detectadas en el archivo: {columnas}")

    texto = []
    texto.append(f"EXCEL: {os.path.basename(file_path)}")
    texto.append("")

    for index, row in df.iterrows():
        fila_texto = []
        for col in columnas:
            valor = row[col]

            # Verificar si el valor es nulo
            if pd.isna(valor):
                valor = "N/A"
            else:
                # Respetar el formato original del valor en Excel
                valor = str(valor)

            fila_texto.append(f"{col.upper()}: {valor}")

        texto.append("\n".join(fila_texto))
        texto.append("")  # Línea en blanco entre filas

    texto_completo = "\n".join(texto)
    download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    temp_output_path = os.path.join(download_folder, "Archivo_Temporal_Ordenado.txt")

    try:
        # Guardar el texto generado en un archivo temporal
        with open(temp_output_path, "w", encoding="utf-8") as file:
            file.write(texto_completo)
        print(f"Archivo temporal generado en: {temp_output_path}")
    except Exception as e:
        print(f"Error al guardar el archivo temporal: {e}")
        return None
    return texto_completo



def identificar_momentos_ai(texto_completo):
    prompt = f"""
    Dado el siguiente texto que contiene operaciones de perforación, realiza lo siguiente:

    1. Analiza el texto completo para identificar las operaciones realizadas.
    2. Identifica palabras clave relevantes que indiquen momentos críticos, como "intento de empacamiento", "torsión", "compresión", "sin circulación", etc, todo lo que indique la identificación del mecanismo de atrapamiento.
    3. Respeta la logica de la secuencia de operaciones de acuerdo a las horas.
    4. Clasifica cada operación en una de las siguientes categorías:
        - ANTES DEL ATRAPAMIENTO: Operaciones realizadas antes de que ocurriera un atrapamiento en la sarta.
        - DESPUÉS DEL ATRAPAMIENTO: Operaciones realizadas después de que se haya identificado que la sarta está atrapada.

    Formato de salida esperado:

    ANTES DEL ATRAPAMIENTO:
    - FECHA: <fecha>
    - HORA INICIO: <hora inicio>
    - HORA FIN: <hora fin>
    - BITÁCORA: <descripción>

    DESPUÉS DEL ATRAPAMIENTO:
    - FECHA: <fecha>
    - HORA INICIO: <hora inicio>
    - HORA FIN: <hora fin>
    - BITÁCORA: <descripción>

    Texto a analizar:
    {texto_completo}

    NOTA:
    1. No añadas comentarios ni explicaciones adicionales fuera del formato solicitado.
    2. Usa estrictamente el formato de salida esperado.
    3. Asegúrate de clasificar todas las operaciones correctamente según las descripciones y las palabras clave detectadas.
    """

    try:
        response = client.chat.completions.create(
            model="nvidia/llama-3.1-nemotron-70b-instruct",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            top_p=1,
            max_tokens=2000,
        )

        # Acceso al contenido de la respuesta
        resultado = response.choices[0].message.content.strip()
        print("Clasificación completada con éxito. Resultado:")
        #print(resultado)

        # Guardar resultado en archivo
        output_path = os.path.join(os.path.expanduser("~"), "Downloads", "Identificación_momentos.txt")
        with open(output_path, "w", encoding="utf-8") as file:
            file.write(resultado)
        print(f"Resultados guardados en: {output_path}")

        return output_path

    except Exception as e:
        print(f"Error al procesar el texto con AI: {e}")
        return None



def llenar_hoja_identificacion(file_path_txt):
    """
    Llena la Hoja de Identificación del Mecanismo de Atrapamiento con base en 
    un archivo clasificado en "Antes" y "Después del Atrapamiento".
    Utiliza la AI para analizar cada sección y asignar valores predefinidos.
    """
    prompt = f"""
    Dado el texto clasificado, realiza lo siguiente:
    
    1. Analiza las operaciones en el texto clasificado.
    2. Responde cada pregunta seleccionando una sola opción, según las condiciones descritas a continuación.
    3. Asigna los valores predefinidos correspondientes para cada mecanismo (Empacamiento o Puenteo, Presión Diferencial, Geometría del Pozo) con base en la respuesta seleccionada.
    4. No puedes asignar valores distintos a los predefinidos, no puedes colocar "No aplica", usa los valores predifinidos.  
    5. Identifica palabras clave relevantes que indiquen momentos críticos, como por ejemplo "intento de empacamiento", 
        "sarta atrapada", "torsión", "compresión", "sin circulación", etc, que indique el momento del atrapamiento.

    ### Condiciones para seleccionar la respuesta:
    - **¿Dirección del movimiento de la TP antes del atrapamiento?**
      - Si la sarta se movió hacia arriba: Selecciona "Hacia arriba".
      - Si la sarta se movió hacia abajo: Selecciona "Hacia abajo".
      - Si no hubo movimiento (estática): Selecciona "Estática".
      Valores predefinidos:
        - **Hacia arriba:** Empacamiento o Puenteo = 2, Presión Diferencial = 0, Geometría del Pozo = 2
        - **Hacia abajo:** Empacamiento o Puenteo = 1, Presión Diferencial = 0, Geometría del Pozo = 2
        - **Estática:** Empacamiento o Puenteo = 2, Presión Diferencial = 2, Geometría del Pozo = 0
    
    - **¿Tipo de movimiento descendente de la sarta después del atrapamiento?**
      - Si la sarta se mueve libremente: Selecciona "Libre".
      - Si la sarta tiene restricción: Selecciona "Con restricción".
      - Si la sarta no puede moverse: Selecciona "Imposible".
      Valores predefinidos:
        - **Libre:** Empacamiento o Puenteo = 0, Presión Diferencial = 0, Geometría del Pozo = 2
        - **Con restricción:** Empacamiento o Puenteo = 1, Presión Diferencial = 0, Geometría del Pozo = 2
        - **Imposible:** Empacamiento o Puenteo = 0, Presión Diferencial = 0, Geometría del Pozo = 0
    
    - **¿Tipo de rotación después del atrapamiento?**
      - Si hay rotación libre: Selecciona "Libre".
      - Si la rotación tiene restricción: Selecciona "Con restricción".
      - Si la rotación es imposible: Selecciona "Imposible".
      Valores predefinidos:
        - **Libre:** Empacamiento o Puenteo = 0, Presión Diferencial = 0, Geometría del Pozo = 2
        - **Con restricción:** Empacamiento o Puenteo = 2, Presión Diferencial = 0, Geometría del Pozo = 2
        - **Imposible:** Empacamiento o Puenteo = 0, Presión Diferencial = 0, Geometría del Pozo = 0
    
    - **¿Tipo de circulación después del atrapamiento?**
      - Si hay circulación libre: Selecciona "Libre".
      - Si la circulación tiene restricción: Selecciona "Con restricción".
      - Si la circulación es imposible: Selecciona "Imposible".
      Valores predefinidos:
        - **Libre:** Empacamiento o Puenteo = 0, Presión Diferencial = 2, Geometría del Pozo = 2
        - **Con restricción:** Empacamiento o Puenteo = 2, Presión Diferencial = 0, Geometría del Pozo = 0
        - **Imposible:** Empacamiento o Puenteo = 2, Presión Diferencial = 0, Geometría del Pozo = 0
        
    ### Formato de salida esperado:
    1. Responde cada pregunta con:
       - Pregunta: <Pregunta>
       - Respuesta seleccionada: <Respuesta>
       - Valores asignados:
         - Empacamiento o Puenteo = <Valor>
         - Presión Diferencial = <Valor>
         - Geometría del Pozo = <Valor>
    2. Calcula y presenta la fila "TOTAL" sumando los valores de cada mecanismo (Empacamiento o Puenteo, Presión Diferencial, Geometría del Pozo).
        
    NOTA:
    - Sigue estrictamente las condiciones mencionadas para seleccionar la respuesta adecuada.
    - Calcula correctamente los totales según los valores asignados.
    - No añadas comentarios adicionales.
    - "Levantó" significa que la sarta puede moverse HACIA ARRIBA.
    - "Trabaja sarta": significa que se está intentando liberar la sarta.
    - "Golpes de martillo descendentes. Sin éxito", el movimiento descentente de la sarta es imposible.
    - "Sin movimiento": significa explícitamente que no hay desplazamiento en la sarta.
    - "Intento liberar": generalmente implica que la sarta está atrapada y no se puede mover.
    - "Estática": se refiere a una sarta que no se desplaza ni hacia arriba ni hacia abajo.
    - "Sobretensión" puede interpretarse como un movimiento de sarta "Con restricción".
    - "Sarta empacada" NO lo asocies con atrapamiento, restricciones o movimientos.

    Asignación de valores por opción:
    Cada opción seleccionada tiene un valor predefinido.
    Los valores numéricos dependen del mecanismo (Empacamiento o Puenteo, Presión Diferencial, Geometría del Pozo).
    ### Texto clasificado:
    {texto_ordenado}        

    """
    try:
        # Llamar a la AI para procesar el texto
        response = client.chat.completions.create(
            model="nvidia/llama-3.1-nemotron-70b-instruct",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,  # Baja creatividad para resultados estructurados
            top_p=1,
            max_tokens=2000,
        )

        # Obtener el contenido procesado por la AI
        resultado = response.choices[0].message.content.strip()
        print("Hoja de Identificación generada con éxito. Resultado:")
        #print(resultado)

        # Guardar la Hoja de Identificación en un archivo de texto
        output_txt_path = os.path.join(os.path.expanduser("~"), "Downloads", "Hoja_Identificacion_AI.txt")
        with open(output_txt_path, "w", encoding="utf-8") as file:
            file.write(resultado)
        print(f"Resultados guardados en: {output_txt_path}")

        return output_txt_path

    except Exception as e:
        print(f"Error al procesar el texto con la AI: {e}")
        return None



def limpiar_valores(valor):
    """
    Elimina caracteres problemáticos y los caracteres '**' de un valor si están presentes.
    """
    if isinstance(valor, str):
        # Reemplaza caracteres problemáticos
        # valor = valor.replace('\u2023', '-')  # Reemplaza el punto de lista con un guion
        # Remueve otros caracteres no imprimibles
        # valor = re.sub(r'[^\x00-\x7F]+', '', valor)  # Elimina caracteres no ASCII
        valor = re.sub(r"\*\*", "", valor).strip()  # Reemplaza '**' con vacío
    return valor



def generar_excel_identificacion_ai(file_path_txt):
    """
    Utiliza la AI para generar una tabla correctamente formateada en Excel
    basada en el archivo de texto generado.
    """
    try:
        # Leer el archivo de texto generado por la AI
        with open(file_path_txt, "r", encoding="utf-8") as file:
            contenido_txt = file.read()
        
        print("Texto leído correctamente. Construyendo prompt para AI...")
        
        # Construir el prompt para la AI
        prompt = f"""
        Dado el siguiente texto generado por el análisis previo:

        {contenido_txt}
        Si el contenido dice 0, coloca 0, no borres ni omitas.
        
        Llena el siguiente formato para una hoja de identificación, respetando las siguientes columnas:
        - Pregunta
        - Empacamiento o Puenteo
        - Presión Diferencial
        - Geometría del Pozo
        Solo quiero mi formato como resultado, no coloques nada adicional.
        
        Cada fila debe corresponder al formato estructurado:
        1. Primera fila: Los encabezados con las columnas mencionadas.
        2. La última fila debe incluir la palabra "TOTAL" con los valores numéricos sumados para cada columna.
        
        ## FORMATO OBLIGATORIO
        
        | Pregunta                                                   | Empacamiento o Puenteo | Presión Diferencial | Geometría del Pozo |
        |------------------------------------------------------------|------------------------|---------------------|--------------------|
        | ¿Dirección del movimiento de la TP antes del atrapamiento? |                        |                     |                    |
        | Hacia arriba                                               |                        |                     |                    |
        | Hacia abajo                                                |                        |                     |                    |
        | Estática                                                   |                        |                     |                    |
        | ¿Tipo de movimiento descendente de la sarta después del atrapamiento? |                         |                     |                    |
        | Libre                                                      |                        |                     |                    |
        | Con restricción                                            |                        |                     |                    |
        | Imposible                                                  |                        |                     |                    |
        | ¿Tipo de rotación después del atrapamiento?                |                        |                     |                    |
        | Libre                                                      |                        |                     |                    |
        | Con restricción                                            |                        |                     |                    |
        | Imposible                                                  |                        |                     |                    |
        | ¿Tipo de circulación después del atrapamiento?             |                        |                     |                    |
        | Libre                                                      |                        |                     |                    |
        | Con restricción                                            |                        |                     |                    |
        | Imposible                                                  |                        |                     |                    |
        | TOTAL                                                      |                        |                     |                    |
        
        ### Notas importantes:
        - Primera columna incluir preguntas y opciones de respuesta.
        - Las opciones seleccionadas deben colocando los valores.
        - Las celdas de las opciones no seleccionadas deben estar vacías.
        - La última fila debe ser `TOTAL`, con la suma de los valores numéricos de las columnas `Empacamiento o Puenteo`, `Presión Diferencial`, y `Geometría del Pozo`.
        - Solo colocar valores predefinos, no "(Estatica)".
        
        Formato tabular, ES OBLIGATORIO QUE ESTE COMPLETO, NO BORRAR NI OMITIR FILAS, devuelve el resultado como una tabla con columnas separadas.
        """

        # Llamar a la AI para procesar el contenido
        response = client.chat.completions.create(
            model="nvidia/llama-3.1-nemotron-70b-instruct",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            top_p=1,
            max_tokens=2000,
        )

        # Obtener el contenido procesado por la AI
        resultado = response.choices[0].message.content.strip()
        print("Tabla generada exitosamente por la AI. Resultado:")
        print(resultado)

        # Crear archivo Excel basado en el resultado procesado
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja de Identificación"

        # Procesar las filas generadas por la AI
        filas = resultado.splitlines()
        for fila in filas:
            # Ignorar líneas vacías o separadores
            if not fila.strip() or "---" in fila:
                continue

            # Dividir la fila por '|' para extraer las columnas
            columnas = [limpiar_valores(col.strip()) for col in fila.split("|")]
            ws.append(columnas)

        # Eliminar la primera columna vacía si existe
        max_col = ws.max_column
        if max_col > 1:  # Verifica que haya más de una columna para eliminar
            ws.delete_cols(1)

        # Aplicar formato a las celdas del encabezado
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Aplicar formato a las celdas del cuerpo
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.value = limpiar_valores(cell.value)  # Limpia los valores en cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        # Guardar el archivo Excel
        output_excel_path = os.path.join(os.path.expanduser("~"), "Downloads", "Hoja_Identificacion_AI.xlsx")
        wb.save(output_excel_path)
        print(f"Archivo Excel generado en: {output_excel_path}")

        return output_excel_path

    except Exception as e:
        # Registrar errores en un archivo de depuración
        debug_log_path = os.path.join(os.path.expanduser("~"), "Downloads", "debug.log")
        with open(debug_log_path, "w", encoding="utf-8") as log_file:
            log_file.write(f"Error detectado en generar_excel_identificacion_ai: {e}\n")
        raise



import sys
if __name__ == "__main__":
    try:
        # Aquí va el flujo principal del script
        file_path = sys.argv[1]  # Leer la ruta del archivo como argumento
        #file_path = r"D:\ESIA Ticoman\ESIA. Semestre 9\Proyecto Terminal 2\PT2. Proyecto en equipo\Pozo 2. Quesqui 35\Quesqui 35. SIOP Filtrado.xlsx"
        
        print("EJECUTANDO AI")
        texto_ordenado = convertir_excel_a_texto(file_path)

        if texto_ordenado:
            archivo_clasificado = identificar_momentos_ai(texto_ordenado)

            if archivo_clasificado:
                print("Paso 3: Ejecutando llenar_hoja_identificacion...")
                archivo_identificacion_txt = llenar_hoja_identificacion(archivo_clasificado)

                if archivo_identificacion_txt and os.path.exists(archivo_identificacion_txt):
                    print("Paso 4: Ejecutando generar_excel_identificacion_ai...")
                    generar_excel_identificacion_ai(archivo_identificacion_txt)
                else:
                    print("Error: Hoja de identificación no generada correctamente.")
            else:
                print("Error: Clasificación de momentos no generada correctamente.")
        else:
            print("Error: No se pudo convertir el archivo Excel a texto.")

    except Exception as e:
        # Captura cualquier error no manejado y lo escribe en un archivo de depuración
        debug_log_path = os.path.join(os.path.expanduser("~"), "Downloads", "debug.log")
        with open(debug_log_path, "w", encoding="utf-8") as log_file:
            log_file.write(f"Error detectado: {e}\n")
        raise  # Opcional: re-lanzar la excepción para depuración en consola